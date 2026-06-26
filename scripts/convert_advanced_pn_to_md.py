import openpyxl
import re
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Advanced PN List.xlsx"
DST_FULL = ROOT / "data" / "converted" / "Advanced PN List.md"
DST_SUMMARY = ROOT / "data" / "converted" / "Advanced PN List Summary.md"

ALL_COLS = list(range(1, 27))

HEADER_LABELS = [
    "Business Unit", "PN", "PN Description", "Product Hierarchy", "Offering Hierarchy Code",
    "Offering Bucket", "OH L1", "OH L2", "OH L3", "OH L4", "OH L5", "OH L6",
    "Product Group", "Sub Product Group", "AI Feature Category", "AI Stage",
    "Vertical Solution L1", "Vertical Solution L2", "Vertical Solution L3",
    "Tower", "Material Type", "SDF", "Announce Date", "State", "Data Source", "Data Source Details",
]

BU_COL = 1
DESC_COL = 3
MAT_TYPE_COL = 21
PROD_GROUP_COL = 13
STATE_COL = 24
OH_L1_COL = 7
OH_L2_COL = 8
DS_COL = 25
DATE_COL = 23
PN_COL = 2


def clean(value):
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "<br>")
    s = s.replace("|", "\\|")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def short(value, n=120):
    s = clean(value)
    if len(s) > n:
        s = s[: n - 1] + "…"
    return s


def pct(n, total):
    if total == 0:
        return "0.0%"
    return f"{n / total * 100:.1f}%"


def render_dist_table(title, counter, total, top_n=None):
    lines = [f"### {title}", ""]
    lines.append(f"| Value | Count | % |")
    lines.append(f"| --- | ---: | ---: |")
    items = counter.most_common(top_n) if top_n else counter.most_common()
    for k, v in items:
        kk = clean(k) if k is not None else "<NULL>"
        if not kk:
            kk = "<NULL>"
        lines.append(f"| {kk} | {v} | {pct(v, total)} |")
    lines.append("")
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Sheet1"]

    rows_out = []
    rows_out.append("| " + " | ".join(HEADER_LABELS) + " |")
    rows_out.append("| " + " | ".join(["---"] * len(ALL_COLS)) + " |")

    bu_counter = Counter()
    mat_type_counter = Counter()
    prod_group_counter = Counter()
    state_counter = Counter()
    oh_l1_counter = Counter()
    oh_l2_counter = Counter()
    ds_counter = Counter()
    samples_by_bu = {}
    dates = []

    count = 0
    for r in range(2, ws.max_row + 1):
        cells = [clean(ws.cell(row=r, column=c).value) for c in ALL_COLS]
        rows_out.append("| " + " | ".join(cells) + " |")
        count += 1

        bu = ws.cell(row=r, column=BU_COL).value
        mt = ws.cell(row=r, column=MAT_TYPE_COL).value
        pg = ws.cell(row=r, column=PROD_GROUP_COL).value
        st = ws.cell(row=r, column=STATE_COL).value
        l1 = ws.cell(row=r, column=OH_L1_COL).value
        l2 = ws.cell(row=r, column=OH_L2_COL).value
        dsrc = ws.cell(row=r, column=DS_COL).value
        dt = ws.cell(row=r, column=DATE_COL).value

        bu_key = bu if bu else "<NULL>"
        bu_counter[bu_key] += 1
        mat_type_counter[mt if mt else "<NULL>"] += 1
        prod_group_counter[pg if pg else "<NULL>"] += 1
        state_counter[st if st else "<NULL>"] += 1
        oh_l1_counter[l1 if l1 else "<NULL>"] += 1
        oh_l2_counter[l2 if l2 else "<NULL>"] += 1
        ds_counter[dsrc if dsrc else "<NULL>"] += 1
        if dt:
            dates.append(str(dt))

        if len(samples_by_bu.get(bu_key, [])) < 3:
            samples_by_bu.setdefault(bu_key, []).append({
                "pn": ws.cell(row=r, column=PN_COL).value,
                "desc": ws.cell(row=r, column=DESC_COL).value,
                "l1": l1,
                "l2": l2,
                "mt": mt,
                "state": st,
            })

    DST_FULL.write_text("\n".join(rows_out), encoding="utf-8")

    summary = []
    summary.append("# Advanced PN List — Summary Report")
    summary.append("")
    summary.append("## 1. 概览")
    summary.append("")
    summary.append(f"- **来源文件**: `data/raw/Advanced PN List.xlsx` (Sheet: `Sheet1`)")
    summary.append(f"- **总行数**: {count:,}")
    summary.append(f"- **总列数**: {len(ALL_COLS)}")
    summary.append(f"- **生成时间**: 一次性快照,重新运行 `scripts/convert_advanced_pn_to_md.py` 可刷新")
    summary.append("")
    summary.append("## 2. 业务单元 (Business Unit) 分布")
    summary.append("")
    bu_total = sum(bu_counter.values())
    summary.append(f"共 {len(bu_counter)} 个 BU,合计 {bu_total:,} 行。")
    summary.append("")
    summary.append(render_dist_table("BU 分布", bu_counter, bu_total))
    summary.append("> 注:本数据集 BU 标签为 `ISG / ISU / MBG / PCSD`,与项目其他数据集中的 `IDG / DCG` 命名体系不同。")
    summary.append("")

    summary.append("## 3. 物料类型 (Material Type) 分布")
    summary.append("")
    summary.append(render_dist_table("Material Type", mat_type_counter, sum(mat_type_counter.values())))
    summary.append("## 4. 产品组 (Product Group) 分布")
    summary.append("")
    summary.append(render_dist_table("Product Group", prod_group_counter, sum(prod_group_counter.values())))
    summary.append("## 5. 状态 (State) 分布")
    summary.append("")
    summary.append(render_dist_table("State", state_counter, sum(state_counter.values())))
    summary.append("## 6. Top 10 OH L1 类别")
    summary.append("")
    summary.append(render_dist_table("OH L1 (Top 10)", oh_l1_counter, sum(oh_l1_counter.values()), top_n=10))
    summary.append("## 7. Top 10 OH L2 类别")
    summary.append("")
    summary.append(render_dist_table("OH L2 (Top 10)", oh_l2_counter, sum(oh_l2_counter.values()), top_n=10))
    summary.append("## 8. 数据来源 (Data Source) 分布")
    summary.append("")
    summary.append(render_dist_table("Data Source", ds_counter, sum(ds_counter.values())))

    summary.append("## 9. Announce Date 范围")
    summary.append("")
    if dates:
        sorted_dates = sorted(dates)
        summary.append(f"- **最早**: `{sorted_dates[0]}`")
        summary.append(f"- **最晚**: `{sorted_dates[-1]}`")
        summary.append(f"- **非空记录数**: {len(dates):,} / {count:,} ({pct(len(dates), count)})")
    else:
        summary.append("- 无有效日期记录")
    summary.append("")

    summary.append("## 10. 各 BU 示例 (每 BU 3 条)")
    summary.append("")
    for bu_key in sorted(samples_by_bu.keys()):
        summary.append(f"### {bu_key}")
        summary.append("")
        summary.append("| PN | Description | OH L1 | OH L2 | Material Type | State |")
        summary.append("| --- | --- | --- | --- | --- | --- |")
        for s in samples_by_bu[bu_key]:
            summary.append(
                f"| `{short(s['pn'], 30)}` | {short(s['desc'], 80)} | {short(s['l1'], 30)} | "
                f"{short(s['l2'], 30)} | {short(s['mt'], 15)} | {short(s['state'], 15)} |"
            )
        summary.append("")

    summary.append("## 11. 数据观察")
    summary.append("")
    summary.append(f"- 数据规模较大 ({count:,} 行),完整转储见 `Advanced PN List.md`。")
    isg_count = bu_counter.get("ISG", 0)
    pcsd_count = bu_counter.get("PCSD", 0)
    big_two = isg_count + pcsd_count
    summary.append(f"- BU 分布呈两极:`ISG` ({isg_count:,}) 与 `PCSD` ({pcsd_count:,}) 合计占比 {pct(big_two, count)},`ISU` 与 `MBG` 体量很小。")
    summary.append(f"- 共 {len(prod_group_counter)} 个 Product Group,共 {len(oh_l1_counter)} 个 OH L1 类别。")
    if dates:
        sorted_dates = sorted(dates)
        if sorted_dates[-1].startswith("9999"):
            summary.append(f"- ⚠️ Announce Date 字段含哨兵值 `9999-12-31`,可能为 '未发布/未填' 占位符,统计时间范围时请剔除。")
    summary.append(f"- 完整字段说明请参考源文件 `Sheet1` 表头。")
    summary.append("")

    DST_SUMMARY.write_text("\n".join(summary), encoding="utf-8")

    print(f"Wrote {DST_FULL}")
    print(f"  Data rows: {count}")
    print(f"  File size: {DST_FULL.stat().st_size:,} bytes")
    print(f"Wrote {DST_SUMMARY}")
    print(f"  File size: {DST_SUMMARY.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
