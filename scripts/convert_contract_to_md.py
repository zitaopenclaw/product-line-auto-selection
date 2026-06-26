import openpyxl
import re
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Contract_Result_20260616-2_UAT.xlsx"
DST_FULL = ROOT / "data" / "converted" / "Contract_Result_20260616-2_UAT.md"
DST_SUMMARY = ROOT / "data" / "converted" / "Contract_Result Summary.md"

COLS = [
    (1, "Contract Number"),
    (2, "Service Model"),
    (3, "Agreement Type"),
    (4, "Business Group"),
    (5, "Created Date"),
    (6, "Opp-ID"),
    (7, "End Customer Country"),
    (8, "Sold-To Country"),
    (9, "Geo"),
    (10, "Sales Org"),
    (11, "Office"),
    (13, "Item No."),
    (14, "Product ID"),
    (19, "Product Category"),
    (21, "Product Type"),
    (35, "Manufacturer"),
    (43, "Contract Price"),
    (44, "Contract Price Currency"),
    (48, "List Price"),
    (51, "Quantity"),
    (52, "Item Start Date"),
    (53, "Item End Date"),
    (55, "Status"),
]

BG_COL = 4
SVC_MODEL_COL = 2
AGREEMENT_COL = 3
CREATED_COL = 5
OPP_COL = 6
END_COUNTRY_COL = 7
GEO_COL = 9
SALES_ORG_COL = 10
CONTRACT_NUM_COL = 1
ITEM_START_COL = 52
ITEM_END_COL = 53
CURRENCY_COL = 44
PRICE_COL = 43
STATUS_COL = 55


def clean(value):
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "<br>")
    s = s.replace("|", "\\|")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def short(value, n=100):
    s = clean(value)
    if len(s) > n:
        s = s[: n - 1] + "…"
    return s


def pct(n, total):
    if total == 0:
        return "0.0%"
    return f"{n / total * 100:.1f}%"


def render_dist_table(title, counter, total, top_n=None):
    lines = [f"### {title}", ""]
    lines.append(f"| Value | Count | % |")
    lines.append(f"| --- | ---: | ---: |")
    items = counter.most_common(top_n) if top_n else counter.most_common()
    for k, v in items:
        kk = clean(k) if k is not None else "<NULL>"
        if not kk:
            kk = "<NULL>"
        lines.append(f"| {kk} | {v} | {pct(v, total)} |")
    lines.append("")
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Result 1"]

    header_labels = [label for _, label in COLS]
    col_indexes = [c for c, _ in COLS]
    rows_out = []
    rows_out.append("| " + " | ".join(header_labels) + " |")
    rows_out.append("| " + " | ".join(["---"] * len(COLS)) + " |")

    bg_counter = Counter()
    svc_model_counter = Counter()
    agreement_counter = Counter()
    geo_counter = Counter()
    end_country_counter = Counter()
    sales_org_counter = Counter()
    status_counter = Counter()
    currency_counter = Counter()
    price_by_currency = {}
    unique_contracts = set()
    unique_opps = set()
    created_dates = []
    item_starts = []
    item_ends = []
    samples = {"IDG": [], "ISG": []}

    count = 0
    for r in range(2, ws.max_row + 1):
        cnum = ws.cell(row=r, column=CONTRACT_NUM_COL).value
        if cnum is None and ws.cell(row=r, column=BG_COL).value is None:
            continue
        cells = [clean(ws.cell(row=r, column=c).value) for c in col_indexes]
        rows_out.append("| " + " | ".join(cells) + " |")
        count += 1

        bg = ws.cell(row=r, column=BG_COL).value
        sm = ws.cell(row=r, column=SVC_MODEL_COL).value
        ag = ws.cell(row=r, column=AGREEMENT_COL).value
        cd = ws.cell(row=r, column=CREATED_COL).value
        opp = ws.cell(row=r, column=OPP_COL).value
        cntry = ws.cell(row=r, column=END_COUNTRY_COL).value
        geo = ws.cell(row=r, column=GEO_COL).value
        so = ws.cell(row=r, column=SALES_ORG_COL).value
        sdt = ws.cell(row=r, column=ITEM_START_COL).value
        edt = ws.cell(row=r, column=ITEM_END_COL).value
        cur = ws.cell(row=r, column=CURRENCY_COL).value
        prc = ws.cell(row=r, column=PRICE_COL).value
        st = ws.cell(row=r, column=STATUS_COL).value

        bg_counter[bg if bg else "<NULL>"] += 1
        svc_model_counter[sm if sm else "<NULL>"] += 1
        agreement_counter[ag if ag else "<NULL>"] += 1
        geo_counter[geo if geo else "<NULL>"] += 1
        end_country_counter[cntry if cntry else "<NULL>"] += 1
        sales_org_counter[so if so else "<NULL>"] += 1
        status_counter[st if st else "<NULL>"] += 1
        currency_counter[cur if cur else "<NULL>"] += 1

        if cnum is not None:
            unique_contracts.add(cnum)
        if opp is not None:
            unique_opps.add(opp)
        if cd:
            created_dates.append(str(cd))
        if sdt:
            item_starts.append(str(sdt))
        if edt:
            item_ends.append(str(edt))

        if isinstance(prc, (int, float)):
            cur_key = cur if cur else "<NULL>"
            price_by_currency[cur_key] = price_by_currency.get(cur_key, 0) + float(prc)

        bg_key = bg if bg in samples else None
        if bg_key and len(samples[bg_key]) < 3:
            samples[bg_key].append({
                "cnum": cnum,
                "bg": bg,
                "sm": sm,
                "ag": ag,
                "opp": opp,
                "cntry": cntry,
                "geo": geo,
                "prc": prc,
                "cur": cur,
                "sdt": sdt,
                "edt": edt,
                "st": st,
            })

    DST_FULL.write_text("\n".join(rows_out), encoding="utf-8")

    summary = []
    summary.append("# Contract_Result — Summary Report")
    summary.append("")
    summary.append("## 1. 概览")
    summary.append("")
    summary.append(f"- **来源文件**: `data/raw/Contract_Result_20260616-2_UAT.xlsx` (Sheet: `Result 1`)")
    summary.append(f"- **数据行数**: {count:,}")
    summary.append(f"- **唯一合同数 (Contract Number)**: {len(unique_contracts):,}")
    summary.append(f"- **唯一机会数 (Opp-ID)**: {len(unique_opps):,}")
    summary.append(f"- **选中列数**: {len(COLS)} (原表 66 列,已筛除 WBS/ECCN/重复字段/内部计费字段)")
    summary.append(f"- **生成时间**: 一次性快照,重新运行 `scripts/convert_contract_to_md.py` 可刷新")
    summary.append("")
    summary.append("## 2. 业务组 (Business Group) 分布")
    summary.append("")
    summary.append(f"共 {len(bg_counter)} 个 BG。")
    summary.append("")
    summary.append(render_dist_table("Business Group", bg_counter, sum(bg_counter.values())))
    summary.append("> 注:本数据集使用 `IDG / ISG` 命名,与项目其他数据集中的 `IDG / DCG` 不完全一致。")
    summary.append("")

    summary.append("## 3. 地理 (Geo) 分布")
    summary.append("")
    summary.append(render_dist_table("Geo", geo_counter, sum(geo_counter.values())))
    summary.append("## 4. 销售组织 (Sales Org) 分布")
    summary.append("")
    summary.append(render_dist_table("Sales Org (Top 10)", sales_org_counter, sum(sales_org_counter.values()), top_n=10))
    summary.append("## 5. Top 10 终端客户国家")
    summary.append("")
    summary.append(render_dist_table("End Customer Country (Top 10)", end_country_counter, sum(end_country_counter.values()), top_n=10))
    summary.append("## 6. 服务模式 (Service Model) 分布")
    summary.append("")
    summary.append(render_dist_table("Service Model", svc_model_counter, sum(svc_model_counter.values())))
    summary.append("## 7. 协议类型 (Agreement Type) 分布")
    summary.append("")
    summary.append(render_dist_table("Agreement Type", agreement_counter, sum(agreement_counter.values())))
    summary.append("## 8. 状态 (Status) 分布")
    summary.append("")
    summary.append(render_dist_table("Status", status_counter, sum(status_counter.values())))

    summary.append("## 9. 按货币汇总的合同金额")
    summary.append("")
    summary.append("| Currency | Total Contract Price | # of Items |")
    summary.append("| --- | ---: | ---: |")
    cur_counts = currency_counter
    for cur, total_price in sorted(price_by_currency.items(), key=lambda x: -x[1]):
        cnt = cur_counts.get(cur, 0)
        summary.append(f"| {cur} | {total_price:,.2f} | {cnt} |")
    summary.append("")

    summary.append("## 10. 时间范围")
    summary.append("")
    if created_dates:
        sc = sorted(created_dates)
        summary.append(f"- **合同 Created Date**: `{sc[0]}` ~ `{sc[-1]}` ({len(created_dates):,} 条)")
    if item_starts:
        ss = sorted(item_starts)
        summary.append(f"- **Item Start Date**: `{ss[0]}` ~ `{ss[-1]}` ({len(item_starts):,} 条)")
    if item_ends:
        se = sorted(item_ends)
        summary.append(f"- **Item End Date**: `{se[0]}` ~ `{se[-1]}` ({len(item_ends):,} 条)")
    summary.append("")

    summary.append("## 11. 示例合同 (各 BG 3 条)")
    summary.append("")
    for bg_key in ["IDG", "ISG"]:
        if bg_key not in samples or not samples[bg_key]:
            continue
        summary.append(f"### {bg_key}")
        summary.append("")
        summary.append("| Contract # | Opp-ID | Service Model | Agreement | Geo | Country | Status | Price | Cur | Item Start | Item End |")
        summary.append("| --- | --- | --- | --- | --- | --- | --- | ---: | --- | --- | --- |")
        for s in samples[bg_key]:
            prc = s["prc"]
            prc_s = f"{prc:,.2f}" if isinstance(prc, (int, float)) else ""
            summary.append(
                f"| {short(s['cnum'], 20)} | {short(s['opp'], 20)} | {short(s['sm'], 10)} | "
                f"{short(s['ag'], 10)} | {short(s['geo'], 5)} | {short(s['cntry'], 5)} | "
                f"{short(s['st'], 10)} | {prc_s} | {short(s['cur'], 5)} | {short(s['sdt'], 12)} | {short(s['edt'], 12)} |"
            )
        summary.append("")

    summary.append("## 12. 数据观察")
    summary.append("")
    summary.append(f"- 共 {count:,} 条 item,归属于 {len(unique_contracts):,} 个合同、{len(unique_opps):,} 个机会 — 平均每合同 {count/len(unique_contracts):.1f} 条 item,每机会 {count/len(unique_opps):.1f} 条 item。")
    summary.append(f"- BG 高度集中于 `IDG` ({pct(bg_counter.get('IDG', 0), count)}),`ISG` 仅 {bg_counter.get('ISG', 0)} 条。")
    summary.append(f"- 完整转储见 `Contract_Result_20260616-2_UAT.md`。")
    summary.append(f"- 源文件中另含 `Query` sheet,内容为生成该报表的 SQL 语句,本报告未纳入。")
    summary.append("")

    DST_SUMMARY.write_text("\n".join(summary), encoding="utf-8")

    print(f"Wrote {DST_FULL}")
    print(f"  Data rows: {count}")
    print(f"  File size: {DST_FULL.stat().st_size:,} bytes")
    print(f"Wrote {DST_SUMMARY}")
    print(f"  File size: {DST_SUMMARY.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
