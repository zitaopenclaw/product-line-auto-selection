import openpyxl
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "OH product in D365.xlsx"
DST = ROOT / "data" / "converted" / "OH product in D365.md"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Product Advanced Find View"]

COLS = [(1, "Product"), (4, "Product Name"), (5, "Product ID"), (14, "Status"), (16, "Business Group")]

def clean(value):
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "<br>")
    s = s.replace("|", "\\|")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()

header_labels = [label for _, label in COLS]
rows_out = []
rows_out.append("| " + " | ".join(header_labels) + " |")
rows_out.append("| " + " | ".join(["---"] * len(COLS)) + " |")

kept = 0
skipped = 0
for r in range(2, ws.max_row + 1):
    status = ws.cell(row=r, column=14).value
    if isinstance(status, str) and status.strip().lower() == "retired":
        skipped += 1
        continue
    cells = [clean(ws.cell(row=r, column=c).value) for c, _ in COLS]
    rows_out.append("| " + " | ".join(cells) + " |")
    kept += 1

DST.write_text("\n".join(rows_out), encoding="utf-8")
print(f"Wrote {DST}")
print(f"Kept rows: {kept}")
print(f"Skipped (Retired): {skipped}")
print(f"File size: {DST.stat().st_size} bytes")
