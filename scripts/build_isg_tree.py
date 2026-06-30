"""Build the ISG HW catalog tree JSON from data/raw/DCG_Product_Catagory_*.xlsx.

Tree: L1 (product type) -> L2 (sub-category) -> L3 (product model, leaf).
L3 nodes are the deepest level; each is stored as a leaf with a single "pns"
entry [{pn: L3_category_id, description: L3_description}] so that
_collect_leaf_descs returns the product model name for embedding.

Source file uses flat "exploded" rows (one row = one complete L1/L2/L3 path).
Duplicates are deduplicated during tree build.

Output: output/isg_pn_tree.json  (same schema as advanced_pn_tree.json)

Note: DCG and ISG are synonymous at Lenovo. ISG is the canonical term here.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = REPO_ROOT / "data" / "raw" / "DCG_Product_Catagory_20260624110842.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "isg_pn_tree.json"
SHEET_NAME = "category"

# 1-indexed column positions
COL_L1_ID = 1
COL_L1_DESC = 2
COL_L2_ID = 3
COL_L2_DESC = 4
COL_L3_ID = 5
COL_L3_DESC = 6
TOTAL_COLS = 6


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return p.parse_args()


def load_rows(input_path: Path) -> list[dict[str, Any]]:
    if not input_path.exists():
        sys.exit(f"Input file not found: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {input_path.name}; "
                 f"available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # skip header
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        raw = list(raw) + [None] * max(0, TOTAL_COLS - len(raw))
        l1_id = str(raw[COL_L1_ID - 1] or "").strip()
        l1_desc = str(raw[COL_L1_DESC - 1] or "").strip()
        l2_id = str(raw[COL_L2_ID - 1] or "").strip()
        l2_desc = str(raw[COL_L2_DESC - 1] or "").strip()
        l3_id = str(raw[COL_L3_ID - 1] or "").strip()
        l3_desc = str(raw[COL_L3_DESC - 1] or "").strip()
        if not l1_id or not l1_desc:
            continue
        out.append({
            "l1_id": l1_id, "l1_desc": l1_desc,
            "l2_id": l2_id, "l2_desc": l2_desc,
            "l3_id": l3_id, "l3_desc": l3_desc,
        })
    return out


def build_tree(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build L1->L2->L3(leaf) tree. L3 stores its description as a single 'pns' entry.

    Deduplicates rows — same L3 ID appearing multiple times is stored once.
    Returns list of L1 nodes.
    """
    tree: list[dict[str, Any]] = []
    by_l1: dict[str, dict[str, Any]] = {}
    by_l2: dict[tuple[str, str], dict[str, Any]] = {}
    seen_l3: set[str] = set()

    for r in rows:
        l1_id, l1_desc = r["l1_id"], r["l1_desc"]
        l2_id, l2_desc = r["l2_id"], r["l2_desc"]
        l3_id, l3_desc = r["l3_id"], r["l3_desc"]

        # L1 node (keyed by ID)
        if l1_id not in by_l1:
            l1_node: dict[str, Any] = {
                "name": l1_desc, "category_id": l1_id,
                "pn_count": 0, "child_count": 0, "children": [],
            }
            by_l1[l1_id] = l1_node
            tree.append(l1_node)

        # L2 node
        l2_key = (l1_id, l2_id)
        if l2_id and l2_key not in by_l2:
            l2_node: dict[str, Any] = {
                "name": l2_desc, "category_id": l2_id,
                "pn_count": 0, "child_count": 0, "children": [],
            }
            by_l2[l2_key] = l2_node
            by_l1[l1_id]["children"].append(l2_node)

        # L3 leaf node (deduplicate by L3 ID)
        if l3_id and l3_id not in seen_l3:
            seen_l3.add(l3_id)
            l3_node: dict[str, Any] = {
                "name": l3_desc, "category_id": l3_id,
                "pn_count": 0, "child_count": 0,
                # Single-entry pns so _collect_leaf_descs returns the product name.
                "pns": [{"pn": l3_id, "description": l3_desc}],
            }
            if l2_id and l2_key in by_l2:
                by_l2[l2_key]["children"].append(l3_node)
            else:
                # No L2: attach directly to L1 (edge case)
                by_l1[l1_id]["children"].append(l3_node)

    return tree


def add_rollups(node: dict[str, Any]) -> int:
    if "pns" in node:
        node["child_count"] = len(node["pns"])
        node["pn_count"] = len(node["pns"])
        return node["pn_count"]
    total = 0
    for child in node.get("children", []):
        total += add_rollups(child)
    node["pn_count"] = total
    node["child_count"] = len(node["children"])
    return total


def sort_tree(node: dict[str, Any]) -> None:
    if "pns" in node:
        return
    node["children"].sort(key=lambda c: c["name"])
    for c in node["children"]:
        sort_tree(c)


def _count_at_depth(tree: list[dict[str, Any]], target: int) -> int:
    count = 0
    def dfs(node, d):
        nonlocal count
        if d == target:
            count += 1
            return
        for child in node.get("children", []):
            dfs(child, d + 1)
    for n in tree:
        dfs(n, 1)
    return count


def build_meta(
    input_path: Path,
    rows: list[dict[str, Any]],
    tree: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "source_file": f"data/raw/{input_path.name}",
        "sheet": SHEET_NAME,
        "source_bg": "ISG",
        "level_columns": ["Level 1 Description", "Level 2 Description",
                          "Level 3 Description"],
        "total_rows": len(rows),
        "total_l1": len(tree),
        "l2_nodes": _count_at_depth(tree, 2),
        "l3_nodes": _count_at_depth(tree, 3),
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }


def write_json(out_path: Path, meta: dict[str, Any], tree: list[dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"meta": meta, "tree": tree}
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        os.replace(tmp, out_path)
    except Exception:
        if tmp.exists():
            tmp.unlink()
        raise
    print(f"Wrote {out_path}", file=sys.stderr)


def main() -> None:
    args = parse_args()
    rows = load_rows(args.input)
    print(f"Loaded {len(rows)} rows from {args.input.name}", file=sys.stderr)

    tree = build_tree(rows)
    for n in tree:
        add_rollups(n)
        sort_tree(n)
    tree.sort(key=lambda n: n["name"])

    meta = build_meta(args.input, rows, tree)
    print(
        f"Tree: {meta['total_l1']} L1 / {meta['l2_nodes']} L2 / "
        f"{meta['l3_nodes']} L3 nodes",
        file=sys.stderr,
    )
    write_json(args.output, meta, tree)


if __name__ == "__main__":
    main()
