"""Build the OH tree JSON from data/raw/Advanced PN List.xlsx.

Single-pass: walk L1..L6 path per row, attach {pn, description} to L6 leaf,
then roll up counts, sort, and serialize to JSON.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = REPO_ROOT / "data" / "raw" / "Advanced PN List.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "advanced_pn_tree.json"

COL_BUSINESS_UNIT = 1   # "Business Unit" — ignored per user decision
COL_PN = 2              # "PN"
COL_PN_DESC = 3         # "PN Description"
COL_LEVEL_START = 7     # "OH L1" begins at column 7 (1-indexed)
COL_LEVEL_END = 12      # "OH L6" ends at column 12 (1-indexed, inclusive)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT,
                   help="Path to Advanced PN List.xlsx")
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                   help="Path to write the tree JSON")
    return p.parse_args()


def load_rows(input_path: Path) -> list[dict[str, Any]]:
    """Read the xlsx and return a list of row dicts.

    Each row dict has keys: pn, description, l1..l6.
    """
    if not input_path.exists():
        sys.exit(f"Input file not found: {input_path}")

    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    if "Sheet1" not in wb.sheetnames:
        sys.exit(f"Sheet 'Sheet1' not found in {input_path.name}")
    ws = wb["Sheet1"]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)

    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        # Pad short rows so indexing doesn't crash; missing cols become None.
        raw_list = list(raw) + [None] * (COL_LEVEL_END - len(raw))
        pn = raw_list[COL_PN - 1]
        if pn is None or (isinstance(pn, str) and pn.strip() == ""):
            sys.exit(f"Row with missing PN encountered (col {COL_PN}); aborting.")
        out.append({
            "pn": str(pn).strip(),
            "description": "" if raw_list[COL_PN_DESC - 1] is None
                           else str(raw_list[COL_PN_DESC - 1]).strip(),
            "l1": raw_list[COL_LEVEL_START - 1] or "",
            "l2": raw_list[COL_LEVEL_START] or "",
            "l3": raw_list[COL_LEVEL_START + 1] or "",
            "l4": raw_list[COL_LEVEL_START + 2] or "",
            "l5": raw_list[COL_LEVEL_START + 3] or "",
            "l6": raw_list[COL_LEVEL_START + 4] or "",
        })
    return out


def build_tree(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Single-pass build of the L1..L6 nested tree.

    Returns a list of L1 nodes. Each intermediate node is:
        {"name": str, "pn_count": 0, "child_count": 0, "children": [...]}
    Each L6 node additionally has a "pns" array:
        {"name": str, "pn_count": 0, "child_count": 0,
         "pns": [{"pn": str, "description": str}, ...]}
    Rollups and sorting are applied by later functions.
    """
    # Level-name → children-dict map for easy "find or create".
    tree: list[dict[str, Any]] = []

    # Path → node lookup so each row's L1..L5 walk is O(1) find-or-create.
    by_path: dict[tuple[str, ...], dict[str, Any]] = {}

    for r in rows:
        levels = (str(r["l1"] or "").strip(),
                  str(r["l2"] or "").strip(),
                  str(r["l3"] or "").strip(),
                  str(r["l4"] or "").strip(),
                  str(r["l5"] or "").strip(),
                  str(r["l6"] or "").strip())

        # Walk L1..L5 creating category nodes as needed.
        for depth in range(5):
            path = levels[: depth + 1]
            node = by_path.get(path)
            if node is None:
                node = {"name": levels[depth], "pn_count": 0,
                        "child_count": 0, "children": []}
                by_path[path] = node
                if depth == 0:
                    tree.append(node)
                else:
                    parent = by_path[levels[:depth]]
                    parent["children"].append(node)

        # Create the L6 leaf (PN container).
        l6_path = levels
        l6_node = by_path.get(l6_path)
        if l6_node is None:
            l6_node = {"name": levels[5], "pn_count": 0,
                       "child_count": 0, "pns": []}
            by_path[l6_path] = l6_node
            parent = by_path[levels[:5]]
            parent["children"].append(l6_node)

        l6_node["pns"].append({"pn": r["pn"], "description": r["description"]})

    return tree


def add_rollups(node: dict[str, Any]) -> int:
    """Recursively set pn_count and child_count on every node. Returns pn_count."""
    if "pns" in node:
        node["child_count"] = len(node["pns"])
        node["pn_count"] = len(node["pns"])
        return node["pn_count"]

    total_pn = 0
    for child in node.get("children", []):
        total_pn += add_rollups(child)
    node["pn_count"] = total_pn
    node["child_count"] = len(node.get("children", []))
    return total_pn


def add_rollups_tree(tree: list[dict[str, Any]]) -> None:
    for n in tree:
        add_rollups(n)


def sort_tree(node: dict[str, Any]) -> None:
    """Recursively sort children by name and pns by pn."""
    if "pns" in node:
        node["pns"].sort(key=lambda p: p["pn"])
        return
    node["children"].sort(key=lambda c: c["name"])
    for c in node.get("children", []):
        sort_tree(c)


def sort_tree_root(tree: list[dict[str, Any]]) -> None:
    tree.sort(key=lambda n: n["name"])
    for n in tree:
        sort_tree(n)


LEVEL_COLUMNS = ["OH L1", "OH L2", "OH L3", "OH L4", "OH L5", "OH L6"]


def build_meta(rows: list[dict[str, Any]]) -> dict[str, Any]:
    empty_l_names = sum(
        1 for r in rows
        for k in ("l1", "l2", "l3", "l4", "l5", "l6")
        if not str(r.get(k) or "").strip()
    )
    empty_descriptions = sum(1 for r in rows if not r["description"])
    return {
        "source_file": "data/raw/Advanced PN List.xlsx",
        "sheet": "Sheet1",
        "level_columns": LEVEL_COLUMNS,
        "total_pns": len(rows),
        "total_l1": 0,  # filled in by main() after tree built
        "empty_l_names": empty_l_names,
        "empty_descriptions": empty_descriptions,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }


def _walk(node):
    if isinstance(node, list):
        for item in node:
            yield from _walk(item)
        return
    yield node
    for c in node.get("children", []):
        yield from _walk(c)


def self_check(tree: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    """Raise ValueError if any invariant is broken. No JSON written if so.

    Uses explicit raises (not assert) so the checks survive `python -O`.
    """
    total_pn = sum(n["pn_count"] for n in tree)
    if total_pn != meta["total_pns"]:
        raise ValueError(
            f"pn_count rollup ({total_pn}) != meta.total_pns ({meta['total_pns']})")

    if len(tree) != 6:
        raise ValueError(f"Expected 6 L1 nodes, got {len(tree)}")

    def depth(n: dict[str, Any]) -> int:
        if "pns" in n:
            return 1
        return 1 + max((depth(c) for c in n.get("children", [])), default=0)
    depths = [depth(n) for n in tree]
    if max(depths) != 6:
        raise ValueError(f"Max tree depth is {max(depths)}, expected 6")

    for n in _walk(tree):
        if "pns" in n:
            if n["child_count"] != len(n["pns"]):
                raise ValueError(
                    f"child_count mismatch at L6 '{n['name']}': "
                    f"{n['child_count']} vs {len(n['pns'])}")
        else:
            if n["child_count"] != len(n.get("children", [])):
                raise ValueError(
                    f"child_count mismatch at '{n['name']}': "
                    f"{n['child_count']} vs {len(n.get('children', []))}")

    seen: set[str] = set()
    for n in _walk(tree):
        for p in n.get("pns", []):
            if p["pn"] in seen:
                raise ValueError(f"Duplicate PN: {p['pn']}")
            seen.add(p["pn"])
    if len(seen) != meta["total_pns"]:
        raise ValueError(
            f"Distinct PNs ({len(seen)}) != meta.total_pns ({meta['total_pns']})")


def write_json(out_path: Path, meta: dict[str, Any], tree: list[dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"meta": meta, "tree": tree}
    # Atomic write: dump to .tmp, then os.replace so a crash mid-write
    # never leaves a half-written advanced_pn_tree.json on disk.
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        with tmp_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, out_path)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink()
        raise
    print(f"Wrote {out_path}", file=sys.stderr)


def main() -> None:
    args = parse_args()
    rows = load_rows(args.input)
    print(f"Loaded {len(rows)} rows from {args.input}", file=sys.stderr)

    tree = build_tree(rows)
    print(f"Tree built: {len(tree)} L1 nodes, "
          f"{_count_l6_leaves(tree)} L6 leaves, "
          f"{_count_attached_pns(tree)} PNs attached", file=sys.stderr)

    add_rollups_tree(tree)
    sort_tree_root(tree)

    meta = build_meta(rows)
    meta["total_l1"] = len(tree)

    self_check(tree, meta)
    print("Self-checks passed", file=sys.stderr)

    if meta["empty_l_names"]:
        print(f"Warning: {meta['empty_l_names']} empty OH level cells "
              f"(of {6 * meta['total_pns']} possible).", file=sys.stderr)
    if meta["empty_descriptions"]:
        print(f"Warning: {meta['empty_descriptions']} PNs with empty description.",
              file=sys.stderr)

    write_json(args.output, meta, tree)


def _count_l6_leaves(tree: list[dict[str, Any]]) -> int:
    n = 0
    stack: list[dict[str, Any]] = list(tree)
    while stack:
        node = stack.pop()
        if "pns" in node:
            n += 1
        else:
            stack.extend(node.get("children", []))
    return n


def _count_attached_pns(tree: list[dict[str, Any]]) -> int:
    n = 0
    stack: list[dict[str, Any]] = list(tree)
    while stack:
        node = stack.pop()
        if "pns" in node:
            n += len(node["pns"])
        else:
            stack.extend(node.get("children", []))
    return n


if __name__ == "__main__":
    main()
