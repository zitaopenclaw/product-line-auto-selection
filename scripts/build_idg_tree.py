"""Build the IDG HW catalog tree JSON from data/raw/IDG_Product Category_HW Related.xlsx.

Tree: L1 (device type) -> L2 (brand) -> L3 (product line) -> L4 (model, leaf).
L5 variants are stored as "pns" at L4 leaves so _collect_leaf_descs gives
embedding context (model variant names). L4 nodes without any L5 get a
synthetic single-entry pns list containing the L4 name itself.

Output: output/idg_pn_tree.json  (same schema as advanced_pn_tree.json)
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = REPO_ROOT / "data" / "raw" / "IDG_Product Category_HW Related.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "idg_pn_tree.json"
SHEET_NAME = "sheet1"

# 1-indexed column positions (matches actual sheet layout)
COL_L1_NAME = 1
COL_L2_NAME = 2
COL_L3_NAME = 3
COL_L4_NAME = 4
COL_L5_NAME = 5
COL_L1_CODE = 6
COL_L2_CODE = 7
COL_L3_CODE = 8
COL_L4_CODE = 9
COL_L5_CODE = 10
COL_IS_DEFAULT = 11
TOTAL_COLS = 11


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return p.parse_args()


def load_rows(input_path: Path) -> list[dict[str, Any]]:
    if not input_path.exists():
        sys.exit(f"Input file not found: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {input_path.name}; "
                 f"available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # skip header row
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        raw = list(raw) + [None] * max(0, TOTAL_COLS - len(raw))
        l1 = str(raw[COL_L1_NAME - 1] or "").strip()
        l2 = str(raw[COL_L2_NAME - 1] or "").strip()
        l3 = str(raw[COL_L3_NAME - 1] or "").strip()
        l4 = str(raw[COL_L4_NAME - 1] or "").strip()
        l5 = str(raw[COL_L5_NAME - 1] or "").strip()
        l4_code = str(raw[COL_L4_CODE - 1] or "").strip()
        l5_code = str(raw[COL_L5_CODE - 1] or "").strip()
        if not l1 or not l2:
            continue
        out.append({
            "l1": l1, "l2": l2, "l3": l3, "l4": l4, "l5": l5,
            "l4_code": l4_code, "l5_code": l5_code,
        })
    return out


def build_tree(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build L1->L2->L3->L4(leaf) tree. L5 variants stored as pns at L4.

    Returns list of L1 nodes.
    """
    tree: list[dict[str, Any]] = []
    by_path: dict[tuple[str, ...], dict[str, Any]] = {}
    # Track seen l5 codes per l4 to avoid duplicate pns entries.
    l4_seen_l5: dict[tuple[str, ...], set[str]] = {}

    for r in rows:
        l1, l2, l3, l4 = r["l1"], r["l2"], r["l3"], r["l4"]

        # Build L1/L2/L3 internal nodes.
        for depth, (level_names, name) in enumerate(
            zip(
                [("l1",), ("l1", "l2"), ("l1", "l2", "l3")],
                [l1, l2, l3],
            )
        ):
            if not name:
                break
            path = tuple(r[k] for k in level_names)
            if path not in by_path:
                node: dict[str, Any] = {
                    "name": name, "pn_count": 0, "child_count": 0, "children": []
                }
                by_path[path] = node
                if depth == 0:
                    tree.append(node)
                else:
                    parent_path = path[:-1]
                    by_path[parent_path]["children"].append(node)

        # Build L4 leaf node.
        if not l4 or not l3:
            continue
        l4_path = (l1, l2, l3, l4)
        if l4_path not in by_path:
            l4_node: dict[str, Any] = {
                "name": l4, "pn_count": 0, "child_count": 0, "pns": []
            }
            by_path[l4_path] = l4_node
            l4_seen_l5[l4_path] = set()
            parent = by_path[(l1, l2, l3)]
            parent["children"].append(l4_node)

        l4_node = by_path[l4_path]
        seen = l4_seen_l5[l4_path]

        if r["l5"]:
            key = r["l5_code"] or r["l5"]
            if key not in seen:
                seen.add(key)
                l4_node["pns"].append({
                    "pn": r["l5_code"] or "",
                    "description": r["l5"],
                })

    # L4 nodes with no L5 variants: add a synthetic entry so embedding text
    # includes at least the model name.
    for path, node in by_path.items():
        if "pns" in node and not node["pns"]:
            node["pns"].append({"pn": "", "description": node["name"]})

    return tree


def add_rollups(node: dict[str, Any]) -> int:
    if "pns" in node:
        node["child_count"] = len(node["pns"])
        node["pn_count"] = len(node["pns"])
        return node["pn_count"]
    total = 0
    for child in node.get("children", []):
        total += add_rollups(child)
    node["pn_count"] = total
    node["child_count"] = len(node["children"])
    return total


def sort_tree(node: dict[str, Any]) -> None:
    if "pns" in node:
        node["pns"].sort(key=lambda p: p["description"])
        return
    node["children"].sort(key=lambda c: c["name"])
    for c in node["children"]:
        sort_tree(c)


def _walk(nodes: list[dict[str, Any]]):
    for n in nodes:
        yield n
        if "children" in n:
            yield from _walk(n["children"])


def count_nodes_at_depth(tree: list[dict[str, Any]], target_depth: int) -> int:
    """Count non-empty nodes at a specific 1-indexed depth."""
    count = 0
    def dfs(node, depth):
        nonlocal count
        if depth == target_depth:
            count += 1
            return
        for child in node.get("children", []):
            dfs(child, depth + 1)
    for n in tree:
        dfs(n, 1)
    return count


def build_meta(
    input_path: Path,
    rows: list[dict[str, Any]],
    tree: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "source_file": f"data/raw/{input_path.name}",
        "sheet": SHEET_NAME,
        "source_bg": "IDG",
        "level_columns": ["Level 1 Name", "Level 2 Name", "Level 3 Name",
                          "Level 4 Name", "Level 5 Name"],
        "total_rows": len(rows),
        "total_l1": len(tree),
        "l2_nodes": count_nodes_at_depth(tree, 2),
        "l3_nodes": count_nodes_at_depth(tree, 3),
        "l4_nodes": count_nodes_at_depth(tree, 4),
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }


def write_json(out_path: Path, meta: dict[str, Any], tree: list[dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"meta": meta, "tree": tree}
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        os.replace(tmp, out_path)
    except Exception:
        if tmp.exists():
            tmp.unlink()
        raise
    print(f"Wrote {out_path}", file=sys.stderr)


def main() -> None:
    args = parse_args()
    rows = load_rows(args.input)
    print(f"Loaded {len(rows)} rows from {args.input.name}", file=sys.stderr)

    tree = build_tree(rows)
    for n in tree:
        add_rollups(n)
        sort_tree(n)
    tree.sort(key=lambda n: n["name"])

    meta = build_meta(args.input, rows, tree)
    print(
        f"Tree: {meta['total_l1']} L1 / {meta['l2_nodes']} L2 / "
        f"{meta['l3_nodes']} L3 / {meta['l4_nodes']} L4 nodes",
        file=sys.stderr,
    )
    write_json(args.output, meta, tree)


if __name__ == "__main__":
    main()
