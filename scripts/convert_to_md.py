import openpyxl
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Approved DER in 2026.xlsx"
DST = ROOT / "data" / "converted" / "Approved DER in 2026.md"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Approved DER in 2026"]

def clean(value):
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "<br>")
    s = s.replace("|", "\\|")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()

rows_out = []
rows_out.append("| Opportunity ID | Business Group | Describe the business problem or challenge |")
rows_out.append("| --- | --- | --- |")

count = 0
for r in range(2, ws.max_row + 1):
    opp = clean(ws.cell(row=r, column=3).value)
    grp = clean(ws.cell(row=r, column=4).value)
    desc = clean(ws.cell(row=r, column=5).value)
    rows_out.append(f"| {opp} | {grp} | {desc} |")
    count += 1

DST.write_text("\n".join(rows_out), encoding="utf-8")
print(f"Wrote {DST}")
print(f"Data rows: {count}")
print(f"File size: {DST.stat().st_size} bytes")
