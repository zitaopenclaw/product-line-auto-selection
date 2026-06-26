"""
LEGACY — OH product list loader (replaced by tree-based pipeline on 2026-06-27).

See meeting notes 2026-06-25, data source switch decision (lines 145, 157).
Kept for rollback / reference only. Not imported by the current pipeline.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass
class OHProduct:
    product_guid: str
    product_name: str
    product_id: str
    status: str
    business_group: str
    parent_product: Optional[str] = None
    solution_category: Optional[str] = None
    solution_sub_category: Optional[str] = None
    iso: Optional[str] = None


_OH_COLS = [
    "(Do Not Modify) Product",
    "Product ID",
    "Product Name",
    "Status",
    "Business Group",
    "Parent Product",
    "ISO",
    "Solution Category",
    "Solution Sub-Category",
]


def load_oh(path: Path, drop_retired: bool = True) -> list[OHProduct]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Product Advanced Find View"]
    col = _header_index(ws, _OH_COLS)

    rows = []
    for r in range(2, ws.max_row + 1):
        status = _clean(ws.cell(row=r, column=col.get("Status", 16)).value)
        if drop_retired and (status or "").lower() == "retired":
            continue
        guid = _clean(ws.cell(row=r, column=col.get("(Do Not Modify) Product", 1)).value)
        pid = _clean(ws.cell(row=r, column=col.get("Product ID", 4)).value)
        name = _clean(ws.cell(row=r, column=col.get("Product Name", 5)).value)
        bg = _clean(ws.cell(row=r, column=col.get("Business Group", 18)).value)
        if not name or not pid or not bg:
            continue
        rows.append(OHProduct(
            product_guid=guid or "",
            product_name=name,
            product_id=pid,
            status=status or "",
            business_group=bg,
            parent_product=_clean(ws.cell(row=r, column=col.get("Parent Product", 9)).value),
            solution_category=_clean(ws.cell(row=r, column=col.get("Solution Category", 11)).value),
            solution_sub_category=_clean(ws.cell(row=r, column=col.get("Solution Sub-Category", 12)).value),
            iso=_clean(ws.cell(row=r, column=col.get("ISO", 10)).value),
        ))
    return rows


def index_oh_by_bg(oh_products: list[OHProduct]) -> dict[str, list[OHProduct]]:
    idx: dict[str, list[OHProduct]] = {}
    for p in oh_products:
        idx.setdefault(p.business_group, []).append(p)
    return idx


def _clean(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _header_index(ws, names: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        key = str(raw).strip()
        for name in names:
            if name.lower() in key.lower():
                idx[name] = col
                break
    return idx
