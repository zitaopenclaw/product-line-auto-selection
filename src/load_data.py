from dataclasses import dataclass, field
from typing import Optional
import openpyxl
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "raw"

DER_PATH = DATA_DIR / "Approved DER in 2026.xlsx"

DER_SHEET = "Approved DER in 2026"


@dataclass
class DERRow:
    opportunity_id: str
    business_group: str
    description: str
    # Structured fields (loaded from DER form columns; None if absent in data)
    service_model: Optional[str] = None          # Service Model (DAAS / PROF & MGD SERVICES / IAAS / …)
    is_existing_expansion: Optional[bool] = None  # Expansion of existing TruScale/managed contract
    is_emerging_tech: Optional[bool] = None       # Emerging / AI technology flag
    is_ars: Optional[bool] = None                 # Asset Recovery Services opportunity
    scope: Optional[str] = None                   # Scope of this opportunity



def _clean(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _yn_to_bool(value) -> Optional[bool]:
    s = _clean(value)
    if s is None:
        return None
    return s.lower() == "yes"


def _header_index(ws, names: list[str]) -> dict[str, int]:
    """Build column-name → 1-based column-index from the header row."""
    idx: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        key = str(raw).strip()
        for name in names:
            if name.lower() in key.lower():
                idx[name] = col
                break
    return idx


# --- DER column name fragments to look up (partial match) ---
_DER_COLS = [
    "Opportunity ID",
    "Business Group",
    "Describe the business problem",
    "Service Model",
    "Is this Opportunity an expansion",
    "Involve The Use of Emerging Technology",
    "Is there an opportunity for Lenovo Asset Recovery",
    "The Scope of This Opportunity",
]


def load_der(path: Path = DER_PATH) -> list[DERRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[DER_SHEET]
    col = _header_index(ws, _DER_COLS)

    rows = []
    for r in range(2, ws.max_row + 1):
        def _get(name: str):
            c = col.get(name)
            return ws.cell(row=r, column=c).value if c else None

        opp = _clean(_get("Opportunity ID"))
        bg = _clean(_get("Business Group"))
        desc = _clean(_get("Describe the business problem"))
        if not opp or not bg:
            continue

        rows.append(DERRow(
            opportunity_id=opp,
            business_group=bg,
            description=desc or "",
            service_model=_clean(_get("Service Model")),
            is_existing_expansion=_yn_to_bool(_get("Is this Opportunity an expansion")),
            is_emerging_tech=_yn_to_bool(_get("Involve The Use of Emerging Technology")),
            is_ars=_yn_to_bool(_get("Is there an opportunity for Lenovo Asset Recovery")),
            scope=_clean(_get("The Scope of This Opportunity")),
        ))
    return rows


def stratified_sample(der_rows: list[DERRow], per_bg: int, seed: int = 42) -> list[DERRow]:
    import random
    rng = random.Random(seed)
    by_bg: dict[str, list[DERRow]] = {}
    for r in der_rows:
        by_bg.setdefault(r.business_group, []).append(r)
    out = []
    for bg, items in by_bg.items():
        k = min(per_bg, len(items))
        out.extend(rng.sample(items, k))
    return out


if __name__ == "__main__":
    der = load_der()
    print(f"DER rows: {len(der)}")
    sample = stratified_sample(der, per_bg=25)
    print(f"Sample size: {len(sample)}")
    # Show structured field coverage
    has_sm = sum(1 for r in der if r.service_model)
    has_ars = sum(1 for r in der if r.is_ars)
    has_ai = sum(1 for r in der if r.is_emerging_tech)
    has_scope = sum(1 for r in der if r.scope)
    print(f"Structured field coverage: service_model={has_sm}, ars={has_ars}, ai={has_ai}, scope={has_scope}")
